"""
Service d'export de rapports PDF et Excel
"""
import io
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

import matplotlib
matplotlib.use('Agg')  # Backend non-GUI
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from io import BytesIO

from plateforme_centralisation.mongo import get_mongo_db


class ExportService:
    """Service pour génération de rapports PDF et Excel"""
    
    def __init__(self):
        self.client, self.db = get_mongo_db()
        self.styles = getSampleStyleSheet()
        self._setup_styles()
    
    def _setup_styles(self):
        """Configure les styles pour les PDFs"""
        # Style titre principal
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        # Style sous-titre
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1e293b'),
            spaceAfter=12,
            spaceBefore=12
        ))
    
    def _get_period_filter(self, period: str = 'all') -> Optional[Dict]:
        """Retourne filtre MongoDB basé sur période"""
        now = datetime.utcnow()
        period_map = {
            '7d': now - timedelta(days=7),
            '30d': now - timedelta(days=30),
            '90d': now - timedelta(days=90),
            '1y': now - timedelta(days=365),
        }
        
        if period in period_map:
            return {"ts": {"$gte": period_map[period].isoformat()}}
        return None
    
    def _create_matplotlib_chart(self, data: List[Dict], chart_type: str = 'line', 
                                  title: str = '', xlabel: str = '', ylabel: str = '') -> BytesIO:
        """Crée un graphique matplotlib et retourne bytes"""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Extraire données
        if not data:
            plt.close()
            return None
        
        dates = [datetime.fromisoformat(d['ts'].replace('Z', '+00:00')) for d in data if 'ts' in d]
        values = [float(d.get('value', 0)) for d in data]
        
        if chart_type == 'line':
            ax.plot(dates, values, marker='o', linewidth=2, markersize=6)
        elif chart_type == 'bar':
            ax.bar(dates, values, width=0.8)
        
        ax.set_title(title, fontsize=16, fontweight='bold')
        ax.set_xlabel(xlabel, fontsize=12)
        ax.set_ylabel(ylabel, fontsize=12)
        ax.grid(True, alpha=0.3)
        
        # Format dates
        if len(dates) > 20:
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        else:
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.xticks(rotation=45, ha='right')
        
        plt.tight_layout()
        
        # Convertir en bytes
        buf = BytesIO()
        plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
        buf.seek(0)
        plt.close()
        
        return buf
    
    # ===== GÉNÉRATION PDF =====
    
    def generate_dashboard_pdf(self, source: str, period: str = '30d') -> BytesIO:
        """
        Génère un rapport PDF pour une source de données
        
        Args:
            source: BRVM, WorldBank, IMF, UN_SDG, AfDB
            period: 7d, 30d, 90d, 1y, all
        
        Returns:
            BytesIO contenant le PDF
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        
        # Titre principal
        title_text = f"Rapport {source.upper()} - {datetime.now().strftime('%d/%m/%Y')}"
        story.append(Paragraph(title_text, self.styles['CustomTitle']))
        story.append(Spacer(1, 0.3*inch))
        
        # Requête MongoDB
        query = {"source": source}
        period_filter = self._get_period_filter(period)
        if period_filter:
            query.update(period_filter)
        
        observations = list(self.db.curated_observations.find(query).sort("ts", -1).limit(500))
        
        # Résumé
        summary_data = [
            ['Source', source],
            ['Période', period],
            ['Observations', len(observations)],
            ['Date génération', datetime.now().strftime('%d/%m/%Y %H:%M')]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.4*inch))
        
        # Grouper par dataset
        datasets = {}
        for obs in observations:
            dataset = obs.get('dataset', 'Unknown')
            if dataset not in datasets:
                datasets[dataset] = []
            datasets[dataset].append(obs)
        
        # Pour chaque dataset, créer graphique + tableau
        for dataset_name, dataset_obs in list(datasets.items())[:5]:  # Max 5 datasets
            story.append(Paragraph(f"Dataset: {dataset_name}", self.styles['CustomHeading']))
            
            # Graphique
            chart_buffer = self._create_matplotlib_chart(
                dataset_obs[:50],
                chart_type='line',
                title=f'Évolution {dataset_name}',
                xlabel='Date',
                ylabel='Valeur'
            )
            
            if chart_buffer:
                img = Image(chart_buffer, width=5*inch, height=3*inch)
                story.append(img)
                story.append(Spacer(1, 0.2*inch))
            
            # Tableau dernières valeurs
            table_data = [['Date', 'Clé', 'Valeur']]
            for obs in dataset_obs[:10]:
                table_data.append([
                    obs.get('ts', '')[:10],
                    obs.get('key', '')[:30],
                    f"{obs.get('value', 0):,.2f}"
                ])
            
            data_table = Table(table_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch])
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            
            story.append(data_table)
            story.append(PageBreak())
        
        # Pied de page avec métadonnées
        footer = Paragraph(
            f"<i>Généré par Plateforme de Centralisation CEDEAO - {datetime.now().strftime('%d/%m/%Y %H:%M')}</i>",
            self.styles['Normal']
        )
        story.append(Spacer(1, 0.5*inch))
        story.append(footer)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer
    
    # ===== GÉNÉRATION EXCEL =====
    
    def generate_dashboard_excel(self, source: str, period: str = '30d') -> BytesIO:
        """
        Génère un rapport Excel avec données et graphiques
        
        Args:
            source: BRVM, WorldBank, IMF, UN_SDG, AfDB
            period: 7d, 30d, 90d, 1y, all
        
        Returns:
            BytesIO contenant le fichier Excel
        """
        buffer = BytesIO()
        wb = Workbook()
        
        # Supprimer feuille par défaut
        wb.remove(wb.active)
        
        # Requête MongoDB
        query = {"source": source}
        period_filter = self._get_period_filter(period)
        if period_filter:
            query.update(period_filter)
        
        observations = list(self.db.curated_observations.find(query).sort("ts", -1).limit(1000))
        
        # Feuille résumé
        summary_sheet = wb.create_sheet("Résumé")
        summary_sheet['A1'] = "Rapport Export"
        summary_sheet['A1'].font = Font(size=16, bold=True, color="3B82F6")
        
        summary_data = [
            ['Source', source],
            ['Période', period],
            ['Total Observations', len(observations)],
            ['Date Génération', datetime.now().strftime('%d/%m/%Y %H:%M')],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, start=3):
            summary_sheet[f'A{row_idx}'] = label
            summary_sheet[f'B{row_idx}'] = value
            summary_sheet[f'A{row_idx}'].font = Font(bold=True)
        
        # Auto-ajuster colonnes
        summary_sheet.column_dimensions['A'].width = 20
        summary_sheet.column_dimensions['B'].width = 30
        
        # Grouper par dataset
        datasets = {}
        for obs in observations:
            dataset = obs.get('dataset', 'Unknown')
            if dataset not in datasets:
                datasets[dataset] = []
            datasets[dataset].append(obs)
        
        # Créer feuille par dataset
        for dataset_name, dataset_obs in list(datasets.items())[:10]:  # Max 10 feuilles
            # Nom feuille valide (max 31 char)
            sheet_name = dataset_name[:31]
            ws = wb.create_sheet(sheet_name)
            
            # En-têtes
            headers = ['Date', 'Clé', 'Valeur', 'Dataset', 'Attributs']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Données
            for row_idx, obs in enumerate(dataset_obs, start=2):
                ws.cell(row=row_idx, column=1, value=obs.get('ts', '')[:10])
                ws.cell(row=row_idx, column=2, value=obs.get('key', ''))
                ws.cell(row=row_idx, column=3, value=float(obs.get('value', 0)))
                ws.cell(row=row_idx, column=4, value=obs.get('dataset', ''))
                ws.cell(row=row_idx, column=5, value=str(obs.get('attrs', {}))[:100])
                
                # Format nombres
                ws.cell(row=row_idx, column=3).number_format = '#,##0.00'
            
            # Auto-ajuster colonnes
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 30
            
            # Graphique si suffisamment de données
            if len(dataset_obs) >= 3:
                chart = LineChart()
                chart.title = f"Évolution {dataset_name[:30]}"
                chart.y_axis.title = "Valeur"
                chart.x_axis.title = "Observations"
                
                data = Reference(ws, min_col=3, min_row=1, max_row=min(len(dataset_obs) + 1, 100))
                chart.add_data(data, titles_from_data=True)
                
                ws.add_chart(chart, f"G2")
        
        # Sauvegarder
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    # ===== EXPORT COMPARAISON =====
    
    def generate_comparison_report(self, countries: List[str], indicators: List[str], 
                                   format: str = 'pdf', period: str = '1y') -> BytesIO:
        """
        Génère rapport de comparaison multi-pays
        
        Args:
            countries: Liste codes pays (ex: ['CIV', 'SEN', 'BEN'])
            indicators: Liste indicateurs (ex: ['SP.POP.TOTL', 'NY.GDP.MKTP.CD'])
            format: 'pdf' ou 'excel'
            period: Période
        
        Returns:
            BytesIO contenant le rapport
        """
        if format == 'pdf':
            return self._generate_comparison_pdf(countries, indicators, period)
        else:
            return self._generate_comparison_excel(countries, indicators, period)
    
    def _generate_comparison_pdf(self, countries: List[str], indicators: List[str], 
                                 period: str) -> BytesIO:
        """Génère PDF de comparaison"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        story = []
        
        title = f"Comparaison Multi-Pays - {datetime.now().strftime('%d/%m/%Y')}"
        story.append(Paragraph(title, self.styles['CustomTitle']))
        story.append(Spacer(1, 0.2*inch))
        
        # Pays et indicateurs
        info_text = f"<b>Pays:</b> {', '.join(countries)}<br/><b>Indicateurs:</b> {len(indicators)}"
        story.append(Paragraph(info_text, self.styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Pour chaque indicateur, créer graphique comparatif
        for indicator in indicators[:5]:  # Max 5 indicateurs
            story.append(Paragraph(f"Indicateur: {indicator}", self.styles['CustomHeading']))
            
            # Requête données
            query = {
                "dataset": indicator,
                "country_code": {"$in": countries}
            }
            period_filter = self._get_period_filter(period)
            if period_filter:
                query.update(period_filter)
            
            observations = list(self.db.curated_observations.find(query).sort("ts", -1).limit(200))
            
            # Graphique comparatif
            if observations:
                # Grouper par pays
                country_data = {}
                for obs in observations:
                    key_parts = obs.get('key', '').split('.')
                    country = key_parts[0] if key_parts else 'Unknown'
                    if country not in country_data:
                        country_data[country] = []
                    country_data[country].append(obs)
                
                # Créer graphique multi-lignes
                fig, ax = plt.subplots(figsize=(10, 5))
                for country, data in country_data.items():
                    dates = [datetime.fromisoformat(d['ts'].replace('Z', '+00:00')) for d in data]
                    values = [float(d.get('value', 0)) for d in data]
                    ax.plot(dates, values, marker='o', label=country, linewidth=2)
                
                ax.set_title(f'Comparaison {indicator}', fontsize=14, fontweight='bold')
                ax.set_xlabel('Date')
                ax.set_ylabel('Valeur')
                ax.legend()
                ax.grid(True, alpha=0.3)
                plt.xticks(rotation=45, ha='right')
                plt.tight_layout()
                
                chart_buf = BytesIO()
                plt.savefig(chart_buf, format='png', dpi=120)
                chart_buf.seek(0)
                plt.close()
                
                img = Image(chart_buf, width=8*inch, height=4*inch)
                story.append(img)
            
            story.append(PageBreak())
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def _generate_comparison_excel(self, countries: List[str], indicators: List[str], 
                                   period: str) -> BytesIO:
        """Génère Excel de comparaison"""
        buffer = BytesIO()
        wb = Workbook()
        wb.remove(wb.active)
        
        # Feuille par indicateur
        for indicator in indicators[:10]:
            sheet_name = indicator[:31]
            ws = wb.create_sheet(sheet_name)
            
            # En-têtes
            headers = ['Pays', 'Date', 'Valeur']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            
            # Données
            query = {
                "dataset": indicator,
                "country_code": {"$in": countries}
            }
            period_filter = self._get_period_filter(period)
            if period_filter:
                query.update(period_filter)
            
            observations = list(self.db.curated_observations.find(query).sort("ts", -1).limit(500))
            
            for row_idx, obs in enumerate(observations, start=2):
                key_parts = obs.get('key', '').split('.')
                country = key_parts[0] if key_parts else 'Unknown'
                
                ws.cell(row=row_idx, column=1, value=country)
                ws.cell(row=row_idx, column=2, value=obs.get('ts', '')[:10])
                ws.cell(row=row_idx, column=3, value=float(obs.get('value', 0)))
                ws.cell(row=row_idx, column=3).number_format = '#,##0.00'
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def __del__(self):
        """Fermer connexion MongoDB"""
        if hasattr(self, 'client'):
            self.client.close()


# Instance singleton
export_service = ExportService()
