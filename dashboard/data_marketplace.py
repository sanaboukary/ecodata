"""
Data Marketplace - Téléchargement de données BRVM + Macro
Interface moderne pour clients avec packages freemium
"""
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.decorators.http import require_http_methods
from plateforme_centralisation.mongo import get_mongo_db
from datetime import datetime, timedelta
import csv
import json
import io


# Liste officielle des 47 actions BRVM
ACTIONS_BRVM_47 = [
    'ABJC', 'BICC', 'BNBC', 'BOAB', 'BOABF', 'BOAC', 'BOAM', 'BOAN', 'BOAS',
    'CABC', 'CBIBF', 'CFAC', 'CIEC', 'ECOC', 'ETIT', 'FTSC', 'NEIC', 'NSBC',
    'NTLC', 'ONTBF', 'ORAC', 'ORGT', 'PALC', 'PHCC', 'PRSC', 'SABC', 'SAFC',
    'SCRC', 'SDCC', 'SDSC', 'SEMC', 'SGBC', 'SHEC', 'SIBC', 'SICB', 'SICC',
    'SIVC', 'SLBC', 'SMBC', 'SNTS', 'SOGC', 'STAC', 'STBC', 'SVOC', 'TTLC',
    'TTLS', 'UNLC', 'UNXC'
]

def get_data_stats():
    """Statistiques des données disponibles"""
    _, db = get_mongo_db()
    
    # Compter observations BRVM dans prices_daily (données brutes quotidiennes)
    brvm_total_obs = db.prices_daily.count_documents({})
    
    stats = {
        'brvm': {
            'name': 'BRVM (Bourse)',
            'actions': ACTIONS_BRVM_47,  # 47 actions officielles
            'total_obs': brvm_total_obs,  # Données quotidiennes brutes
            'latest_date': None,
            'attributes': ['Open', 'High', 'Low', 'Close', 'Volume']  # Données OHLCV brutes
        },
        'worldbank': {
            'name': 'Banque Mondiale',
            'indicators': 35,
            'countries': 15,
            'total_obs': db.curated_observations.count_documents({'source': 'WorldBank'}),
            'attributes': ['PIB', 'Population', 'Inflation', 'Croissance', 'Commerce']
        },
        'imf': {
            'name': 'FMI (Fonds Monétaire)',
            'series': 20,
            'total_obs': db.curated_observations.count_documents({'source': 'IMF'}),
            'attributes': ['CPI', 'GDP', 'Taux change', 'Réserves', 'Balance']
        },
        'un_sdg': {
            'name': 'ONU Objectifs Développement',
            'series': 8,
            'total_obs': db.curated_observations.count_documents({'source': 'UN_SDG'}),
            'attributes': ['Pauvreté', 'Éducation', 'Santé', 'Emploi']
        },
        'afdb': {
            'name': 'Banque Africaine Développement',
            'indicators': 6,
            'total_obs': db.curated_observations.count_documents({'source': 'AfDB'}),
            'attributes': ['Dette', 'IDE', 'Balance commerciale']
        },
        'publications': {
            'name': 'Publications BRVM',
            'total_obs': db.curated_observations.count_documents({
                'source': 'BRVM',
                'dataset': {'$in': ['PUBLICATION', 'COMMUNIQUE', 'RAPPORT']}
            }),
            'types': ['Communiqués', 'Rapports', 'Bulletins', 'Résultats']
        }
    }
    
    # Dernière date BRVM depuis prices_daily
    latest_brvm = db.prices_daily.find_one(
        {},
        sort=[('date', -1)]
    )
    if latest_brvm:
        stats['brvm']['latest_date'] = latest_brvm.get('date', '')[:10]
    
    # Total général
    stats['total'] = sum(s.get('total_obs', 0) for s in stats.values() if isinstance(s, dict))
    
    return stats


@require_http_methods(["GET"])
def data_marketplace_page(request):
    """Page principale du marketplace de données"""
    stats = get_data_stats()
    
    # Packages disponibles
    packages = {
        'free': {
            'name': 'Gratuit',
            'price': 0,
            'features': [
                '7 derniers jours BRVM',
                '3 recommandations/jour',
                'Export CSV uniquement',
                'Alertes email quotidiennes',
                'Support communautaire'
            ],
            'limits': {
                'period': '7d',
                'format': ['csv'],
                'downloads': '10/mois'
            }
        },
        'pro': {
            'name': 'Pro',
            'price': 5000,
            'features': [
                '60 jours données complètes',
                'Recommandations IA illimitées',
                'Export CSV/JSON/Excel',
                'Alertes push temps réel',
                'API 1000 req/jour',
                '5 stratégies trading',
                'Support email prioritaire'
            ],
            'limits': {
                'period': '60d',
                'format': ['csv', 'json', 'excel'],
                'downloads': 'illimité'
            }
        },
        'premium': {
            'name': 'Premium',
            'price': 15000,
            'features': [
                'Historique complet (5 ans)',
                'API accès illimité',
                '10 stratégies + backtesting',
                'Alertes SMS',
                'Publications avec NLP',
                'Données macro-économiques',
                'Support téléphone prioritaire',
                'Webinaires exclusifs'
            ],
            'limits': {
                'period': 'all',
                'format': ['csv', 'json', 'excel', 'parquet'],
                'downloads': 'illimité'
            }
        },
        'enterprise': {
            'name': 'Enterprise',
            'price': 50000,
            'features': [
                'API haute fréquence',
                'Données brutes MongoDB',
                'Webhook temps réel',
                'Intégration custom',
                'Support dédié 24/7',
                'Formations sur-mesure',
                'SLA 99.9%',
                'Consultation mensuelle expert'
            ],
            'limits': {
                'period': 'all',
                'format': ['csv', 'json', 'excel', 'parquet', 'sql'],
                'downloads': 'illimité',
                'api': 'illimité'
            }
        }
    }
    
    context = {
        'stats': stats,
        'packages': packages,
        'page_title': 'Marketplace de Données'
    }
    
    return render(request, 'dashboard/data_marketplace.html', context)


@require_http_methods(["POST"])
def prepare_download(request):
    """Préparer le téléchargement avec preview"""
    import json
    
    try:
        data = json.loads(request.body)
        source_input = data.get('source')
        period = data.get('period', '7d')
        format_type = data.get('format', 'csv')
        
        # Mapping sources interface → MongoDB
        source_mapping = {
            'brvm': 'BRVM',
            'worldbank': 'WorldBank',
            'imf': 'IMF',
            'un_sdg': 'UN_SDG',
            'afdb': 'AfDB',
            'publications': 'BRVM_PUBLICATION'
        }
        
        source = source_mapping.get(source_input.lower(), source_input)
        
        _, db = get_mongo_db()
        
        # Calculer date limite selon période
        period_map = {
            '7d': 7,
            '30d': 30,
            '60d': 60,
            '1y': 365,
            'all': None
        }
        
        # 🔥 BRVM: Utiliser prices_daily (données brutes quotidiennes)
        if source == 'BRVM':
            query = {}
            
            if period != 'all':
                days = period_map.get(period, 7)
                threshold = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
                query['date'] = {'$gte': threshold}
            
            # Compter observations depuis prices_daily
            count = db.prices_daily.count_documents(query)
            
            # Preview échantillon
            sample = list(db.prices_daily.find(query).limit(5))
            
            # Formater échantillon
            preview = []
            for doc in sample:
                preview.append({
                    'symbol': doc.get('symbol'),
                    'date': doc.get('date', '')[:10],
                    'close': doc.get('close'),
                    'volume': doc.get('volume')
                })
        else:
            # Autres sources: curated_observations
            query = {'source': source}
            
            if period != 'all':
                days = period_map.get(period, 7)
                threshold = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
                query['ts'] = {'$gte': threshold}
            
            # Compter observations
            count = db.curated_observations.count_documents(query)
            
            # Preview échantillon
            sample = list(db.curated_observations.find(query).limit(5))
            
            # Formater échantillon
            preview = []
            for obs in sample:
                preview.append({
                'key': obs.get('key'),
                'date': obs.get('ts', '')[:10],
                'value': obs.get('value'),
                'attrs': len(obs.get('attrs', {}))
            })
        
        # Estimer taille fichier
        avg_size = 500  # bytes par observation
        estimated_size_mb = (count * avg_size) / (1024 * 1024)
        
        return JsonResponse({
            'success': True,
            'count': count,
            'preview': preview,
            'estimated_size': f"{estimated_size_mb:.2f} MB",
            'format': format_type
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@require_http_methods(["GET"])
def download_data(request):
    """Télécharger les données dans le format demandé"""
    source_input = request.GET.get('source')
    period = request.GET.get('period', '7d')
    format_type = request.GET.get('format', 'csv')
    
    if not source_input:
        return JsonResponse({'error': 'Source manquante'}, status=400)
    
    # Mapping sources interface → MongoDB
    source_mapping = {
        'brvm': 'BRVM',
        'worldbank': 'WorldBank',
        'imf': 'IMF',
        'un_sdg': 'UN_SDG',
        'afdb': 'AfDB',
        'publications': 'BRVM_PUBLICATION'
    }
    
    source = source_mapping.get(source_input.lower(), source_input)
    
    _, db = get_mongo_db()
    
    # Calculer période
    period_map = {
        '7d': 7,
        '30d': 30,
        '60d': 60,
        '1y': 365,
        'all': None
    }
    
    # 🔥 BRVM: Utiliser prices_daily (données brutes quotidiennes)
    if source == 'BRVM':
        query = {}
        
        if period != 'all':
            days = period_map.get(period, 7)
            threshold = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            query['date'] = {'$gte': threshold}
        
        # Récupérer depuis prices_daily
        data = list(db.prices_daily.find(query).sort('date', -1))
    else:
        # Autres sources: curated_observations
        query = {'source': source}
        
        if period != 'all':
            days = period_map.get(period, 7)
            threshold = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            query['ts'] = {'$gte': threshold}
        
        # Récupérer données
        data = list(db.curated_observations.find(query).sort('ts', -1))
    
    # Préparer nom fichier
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename_base = f"{source}_{period}_{timestamp}"
    
    # Export selon format
    if format_type == 'csv':
        return export_csv(data, filename_base, source)
    elif format_type == 'json':
        return export_json(data, filename_base, source)
    elif format_type == 'excel':
        return export_excel(data, filename_base, source)
    else:
        return JsonResponse({'error': 'Format non supporté'}, status=400)


def export_csv(data, filename, source=None):
    """Export en CSV optimisé pour analyses"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    # BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    if not data:
        writer.writerow(['Aucune donnée disponible'])
        return response
    
    # Identifier type de source
    if not source:
        source = data[0].get('source', '')
    
    # 🔥 BRVM: En-têtes données OHLCV brutes (prices_daily)
    if source == 'BRVM':
        headers = [
            'Date', 'Symbol', 'Open', 'High', 'Low', 'Close', 'Volume'
        ]
        writer.writerow(headers)
        
        # Données brutes depuis prices_daily
        for doc in data:
            row = [
                doc.get('date', ''),
                doc.get('symbol', ''),
                doc.get('open', ''),
                doc.get('high', ''),
                doc.get('low', ''),
                doc.get('close', ''),
                doc.get('volume', '')
            ]
            writer.writerow(row)
    
    # Autres sources: logique existante
    elif source in ['WorldBank', 'IMF', 'UN_SDG', 'AfDB']:
        headers = [
            'Date', 'Pays', 'Code_Pays', 'Indicateur', 'Code_Indicateur', 
            'Valeur', 'Année', 'Région'
        ]
        writer.writerow(headers)
        
        for obs in data:
            attrs = obs.get('attrs', {})
            year = obs.get('ts', '')[:4]
            row = [
                obs.get('ts', '')[:10],
                attrs.get('country_name', attrs.get('country', obs.get('key', ''))),
                obs.get('key', ''),
                attrs.get('indicator_name', attrs.get('series_name', obs.get('dataset', ''))),
                obs.get('dataset', ''),
                obs.get('value', ''),
                year,
                attrs.get('region', attrs.get('area', ''))
            ]
            writer.writerow(row)
    
    elif source == 'BRVM_PUBLICATION':
        headers = [
            'Date', 'Type', 'Titre', 'Société', 'Contenu', 'URL'
        ]
        writer.writerow(headers)
        
        for obs in data:
            attrs = obs.get('attrs', {})
            row = [
                obs.get('ts', '')[:10],
                attrs.get('type', obs.get('dataset', '')),
                attrs.get('titre', ''),
                attrs.get('societe', attrs.get('symbole', '')),
                attrs.get('contenu', '')[:200],
                attrs.get('url', '')
            ]
            writer.writerow(row)
    
    else:
        # Format générique
        headers = ['Date', 'Entité', 'Indicateur', 'Valeur']
        all_attrs = set()
        for obs in data:
            if 'attrs' in obs:
                all_attrs.update(obs['attrs'].keys())
        headers.extend(sorted(all_attrs))
        writer.writerow(headers)
        
        for obs in data:
            attrs = obs.get('attrs', {})
            row = [
                obs.get('ts', '')[:10],
                obs.get('key', ''),
                obs.get('dataset', ''),
                obs.get('value', '')
            ]
            for attr in sorted(all_attrs):
                row.append(attrs.get(attr, ''))
            writer.writerow(row)
    
    return response


def export_json(data, filename, source=None):
    """Export en JSON"""
    # Identifier source
    if not source:
        source = data[0].get('source', '') if data else ''
    
    # Nettoyer données pour JSON
    clean_data = []
    
    if source == 'BRVM':
        # Données brutes prices_daily
        for doc in data:
            clean_obs = {
                'date': doc.get('date'),
                'symbol': doc.get('symbol'),
                'open': doc.get('open'),
                'high': doc.get('high'),
                'low': doc.get('low'),
                'close': doc.get('close'),
                'volume': doc.get('volume')
            }
            clean_data.append(clean_obs)
    else:
        # Autres sources
        for obs in data:
            clean_obs = {
                'source': obs.get('source'),
                'dataset': obs.get('dataset'),
                'key': obs.get('key'),
                'date': obs.get('ts', '')[:10],
                'value': obs.get('value'),
                'attributes': obs.get('attrs', {})
            }
            clean_data.append(clean_obs)
    
    response = HttpResponse(
        json.dumps(clean_data, indent=2, ensure_ascii=False, default=str),
        content_type='application/json; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
    
    return response


def export_excel(data, filename, source=None):
    """Export en Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Données"
        
        # Style en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Identifier source
        if not source:
            source = data[0].get('source', '') if data else ''
        
        # 🔥 BRVM: En-têtes données OHLCV brutes
        if source == 'BRVM':
            headers = ['Date', 'Symbol', 'Open', 'High', 'Low', 'Close', 'Volume']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Données depuis prices_daily
            for row_idx, doc in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=doc.get('date', ''))
                ws.cell(row=row_idx, column=2, value=doc.get('symbol', ''))
                ws.cell(row=row_idx, column=3, value=doc.get('open', ''))
                ws.cell(row=row_idx, column=4, value=doc.get('high', ''))
                ws.cell(row=row_idx, column=5, value=doc.get('low', ''))
                ws.cell(row=row_idx, column=6, value=doc.get('close', ''))
                ws.cell(row=row_idx, column=7, value=doc.get('volume', ''))
        
        else:
            # Autres sources: logique existante
            headers = ['Source', 'Dataset', 'Key', 'Date', 'Value']
            all_attrs = set()
            for obs in data:
                if 'attrs' in obs:
                    all_attrs.update(obs['attrs'].keys())
            headers.extend(sorted(all_attrs))
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Données
            for row_idx, obs in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=obs.get('source', ''))
                ws.cell(row=row_idx, column=2, value=obs.get('dataset', ''))
                ws.cell(row=row_idx, column=3, value=obs.get('key', ''))
                ws.cell(row=row_idx, column=4, value=obs.get('ts', '')[:10])
                ws.cell(row=row_idx, column=5, value=obs.get('value', ''))
                
                attrs = obs.get('attrs', {})
                for col_idx, attr in enumerate(sorted(all_attrs), 6):
                    ws.cell(row=row_idx, column=col_idx, value=attrs.get(attr, ''))
        
        # Ajuster largeur colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder dans buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        return response
        
    except ImportError:
        return HttpResponse(
            "openpyxl non installé. Utilisez: pip install openpyxl",
            status=500
        )


@require_http_methods(["GET"])
def api_documentation(request):
    """Documentation de l'API publique"""
    docs = {
        'version': '1.0',
        'base_url': 'https://api.votre-plateforme.com/v1',
        'authentication': {
            'type': 'Bearer Token',
            'header': 'Authorization: Bearer YOUR_API_KEY',
            'get_key': 'Compte > Paramètres > API Keys'
        },
        'endpoints': [
            {
                'path': '/brvm/stocks',
                'method': 'GET',
                'description': 'Liste toutes les actions BRVM avec derniers prix',
                'parameters': {
                    'period': '7d|30d|60d|1y|all (défaut: 7d)',
                    'limit': 'nombre max résultats (défaut: 100)'
                },
                'example': 'GET /brvm/stocks?period=30d&limit=50',
                'response': {
                    'results': [
                        {
                            'symbol': 'BOABF',
                            'date': '2026-01-07',
                            'close': 4563.17,
                            'volume': 40394,
                            'rsi': 70.0
                        }
                    ],
                    'count': 47
                }
            },
            {
                'path': '/brvm/stock/{symbol}',
                'method': 'GET',
                'description': 'Historique détaillé d\'une action',
                'parameters': {
                    'symbol': 'Code action (ex: BOABF)',
                    'period': '7d|30d|60d|1y|all'
                },
                'example': 'GET /brvm/stock/BOABF?period=60d'
            },
            {
                'path': '/brvm/recommendations',
                'method': 'GET',
                'description': 'Recommandations IA du jour',
                'response': {
                    'buy_signals': [],
                    'sell_signals': [],
                    'premium_opportunities': []
                }
            },
            {
                'path': '/worldbank/indicators',
                'method': 'GET',
                'description': 'Indicateurs Banque Mondiale',
                'parameters': {
                    'country': 'Code pays (CIV, SEN, BEN...)',
                    'indicator': 'Code indicateur'
                }
            },
            {
                'path': '/download',
                'method': 'GET',
                'description': 'Télécharger données en masse',
                'parameters': {
                    'source': 'BRVM|WorldBank|IMF|UN_SDG|AfDB',
                    'period': '7d|30d|60d|1y|all',
                    'format': 'csv|json|excel'
                },
                'example': 'GET /download?source=BRVM&period=30d&format=csv'
            }
        ],
        'rate_limits': {
            'free': '100 requêtes/jour',
            'pro': '1,000 requêtes/jour',
            'premium': 'Illimité',
            'enterprise': 'Illimité + haute fréquence'
        },
        'code_examples': {
            'python': '''
import requests

API_KEY = "votre_cle_api"
headers = {"Authorization": f"Bearer {API_KEY}"}

# Récupérer actions BRVM
response = requests.get(
    "https://api.votre-plateforme.com/v1/brvm/stocks",
    headers=headers,
    params={"period": "30d"}
)
data = response.json()

for stock in data["results"]:
    print(f"{stock['symbol']}: {stock['close']} FCFA")
''',
            'javascript': '''
const API_KEY = "votre_cle_api";

fetch("https://api.votre-plateforme.com/v1/brvm/stocks?period=30d", {
    headers: {
        "Authorization": `Bearer ${API_KEY}`
    }
})
.then(res => res.json())
.then(data => {
    data.results.forEach(stock => {
        console.log(`${stock.symbol}: ${stock.close} FCFA`);
    });
});
''',
            'r': '''
library(httr)
library(jsonlite)

API_KEY <- "votre_cle_api"

response <- GET(
    "https://api.votre-plateforme.com/v1/brvm/stocks",
    add_headers(Authorization = paste("Bearer", API_KEY)),
    query = list(period = "30d")
)

data <- fromJSON(content(response, "text"))
print(data$results)
'''
        }
    }
    
    return JsonResponse(docs, json_dumps_params={'indent': 2})


@require_http_methods(["GET"])
def get_available_years(request):
    """Récupérer années disponibles par source"""
    _, db = get_mongo_db()
    
    source = request.GET.get('source', '')
    
    # Mapping
    source_mapping = {
        'brvm': 'BRVM',
        'worldbank': 'WorldBank',
        'imf': 'IMF',
        'un_sdg': 'UN_SDG',
        'afdb': 'AfDB',
        'publications': 'BRVM_PUBLICATION'
    }
    
    source_mapped = source_mapping.get(source.lower(), source)
    
    # Récupérer années distinctes
    years_from_attrs = db.curated_observations.distinct('attrs.year', {
        'source': source_mapped,
        'attrs.year': {'$exists': True, '$ne': None}
    })
    
    # Extraire années depuis champ ts
    pipeline = [
        {'$match': {'source': source_mapped}},
        {'$project': {
            'year': {'$substr': ['$ts', 0, 4]}
        }},
        {'$group': {'_id': '$year'}},
        {'$sort': {'_id': -1}}
    ]
    
    years_from_ts = [int(doc['_id']) for doc in db.curated_observations.aggregate(pipeline)]
    
    # Fusionner
    all_years = sorted(set(years_from_attrs + years_from_ts), reverse=True)
    
    return JsonResponse({
        'years': all_years,
        'source': source_mapped
    })


@require_http_methods(["GET"])
def get_available_datasets(request):
    """Récupérer datasets disponibles par source"""
    _, db = get_mongo_db()
    
    source = request.GET.get('source', '')
    year = request.GET.get('year', None)
    
    # Mapping
    source_mapping = {
        'brvm': 'BRVM',
        'worldbank': 'WorldBank',
        'imf': 'IMF',
        'un_sdg': 'UN_SDG',
        'afdb': 'AfDB',
        'publications': 'BRVM_PUBLICATION'
    }
    
    source_mapped = source_mapping.get(source.lower(), source)
    
    # Query
    query = {'source': source_mapped}
    
    if year and year != 'all':
        # Filtrer par année
        query['$or'] = [
            {'attrs.year': int(year)},
            {'ts': {'$regex': f'^{year}'}}
        ]
    
    # Datasets distincts avec infos
    pipeline = [
        {'$match': query},
        {'$group': {
            '_id': '$dataset',
            'count': {'$sum': 1},
            'first_date': {'$min': '$ts'},
            'last_date': {'$max': '$ts'},
            'sample_key': {'$first': '$key'},
            'sample_value': {'$first': '$value'},
            'indicator_name': {'$first': '$attrs.indicator_name'},
            'series_name': {'$first': '$attrs.series_name'}
        }},
        {'$sort': {'count': -1}}
    ]
    
    datasets_cursor = db.curated_observations.aggregate(pipeline)
    
    datasets = []
    for doc in datasets_cursor:
        name = doc.get('indicator_name') or doc.get('series_name') or doc['_id']
        datasets.append({
            'code': doc['_id'],
            'name': name,
            'count': doc['count'],
            'first_date': doc['first_date'][:10] if doc.get('first_date') else '',
            'last_date': doc['last_date'][:10] if doc.get('last_date') else '',
            'sample_key': doc.get('sample_key', ''),
            'sample_value': doc.get('sample_value', '')
        })
    
    return JsonResponse({
        'datasets': datasets,
        'source': source_mapped,
        'year': year,
        'total': len(datasets)
    })
